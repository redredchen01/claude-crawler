import os
import json
from flask import Flask, render_template, request, jsonify
from openpyxl import load_workbook
from werkzeug.utils import secure_filename

app = Flask(__name__)

# Configuration
UPLOAD_FOLDER = 'uploads'
ALLOWED_EXTENSIONS = {'xlsx', 'xls'}
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER

# Ensure upload directory exists
if not os.path.exists(UPLOAD_FOLDER):
    os.makedirs(UPLOAD_FOLDER)

def allowed_file(filename):
    return '.' in filename and \
           filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/api/upload', methods=['POST'])
def upload_file():
    if 'file' not in request.files:
        return jsonify({'error': 'No file part'}), 400
    file = request.files['file']
    if file.filename == '':
        return jsonify({'error': 'No selected file'}), 400
    if file and allowed_file(file.filename):
        filename = secure_filename(file.filename)
        filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        file.save(filepath)
        
        # Parse Excel
        try:
            wb = load_workbook(filepath, data_only=True)
            sheet = wb.active
            rows = []
            for row in sheet.iter_rows(values_only=True):
                # Filter out empty rows
                if any(row):
                    rows.append(list(row))
            
            return jsonify({
                'filename': filename,
                'data': rows
            })
        except Exception as e:
            return jsonify({'error': f'Failed to parse Excel: {str(e)}'}), 500
    return jsonify({'error': 'Invalid file type. Only Excel files are allowed.'}), 400

@app.route('/api/send', methods=['POST'])
def send_emails():
    # Simulate email sending logic
    data = request.json
    emails_to_send = data.get('emails', [])
    
    # In a real app, you would iterate and use an SMTP server
    # For now, we simulate a small delay and success
    import time
    time.sleep(1) 
    
    return jsonify({'message': f'Successfully simulated sending {len(emails_to_send)} emails!'})

if __name__ == '__main__':
    # Default port set to 8787 as per user's tutorial requirement
    app.run(host='0.0.0.0', port=8787, debug=True)
